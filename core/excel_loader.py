"""=================不积跬步无以至千里==================
作    者： 莫景怀
创建时间： 2023/5/9 23:39
文件名： excel_public.py
功能作用：
=================不积小流无以成江海=================="""
import openpyxl
from tools import config
import os
import warnings
#没有默认样式的excel文档是由java程序生成的时候，对文件读写会警告。此行忽略警告级别的提示
warnings.filterwarnings('ignore')
from tools.config import DATA_DIR


class ExcelFile:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = openpyxl.load_workbook(self.filename)


    def get_sheet_names(self):
        return self.workbook.sheetnames

    def get_sheet_by_name(self, sheet_name):
        return self.workbook[sheet_name]

    def create_sheet(self, sheet_name):
        self.workbook.create_sheet(sheet_name)
        self.save()


    # 删除sheet
    def remove_sheet(self, sheet_name):
        sheet = self.workbook[sheet_name]
        self.workbook.remove(sheet)

    def get_all_rows(self, sheet_name):
        sheet = self.workbook[sheet_name]
        return sheet.values

    def get_row(self, sheet_name, row_num):
        sheet = self.workbook[sheet_name]
        return sheet[row_num]

    # 读取指定单元格的值
    def get_cell(self, sheet_name, row_num, col_num):
        sheet = self.workbook[sheet_name]
        return sheet.cell(row=row_num, column=col_num).value

    # 写入指定单元格
    def set_cell_value(self, sheet_name, row_num, col_num, value):
        sheet = self.workbook[sheet_name]
        sheet.cell(row=row_num, column=col_num, value=value)

    # 写入之后保存
    def save(self):
        self.workbook.save(self.filename)

    # 获取指定列，返回集合，其中去除第一行
    def column_values_to_set(self, sheet_name, col_num):
        sheet = self.workbook[sheet_name]
        values = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # 如果为None，跳过此次
            if row[col_num - 1] is None:
                continue
            values.add(row[col_num - 1])
        # print(len(values))
        return values


    # 获取指定两列，生成字典，去除第一行
    def columns_values_to_dict(self, sheet_name, col_num1, col_num2):
        sheet = self.workbook[sheet_name]
        values = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # 如果为None，跳过此次
            if row[col_num1 - 1] is None or row[col_num2 - 1] is None:
                continue
            values[row[col_num1 - 1]] = row[col_num2 - 1]
        return values

    # 获取指定两行，组成字典
    def rows_values_to_dict(self, sheet_name, row_num1, row_num2):
        sheet = self.workbook[sheet_name]
        values = {}
        for col in sheet.iter_cols(min_row=1, values_only=True):
            # 如果为None，跳过此次
            if col[row_num1 - 1] is None or col[row_num2 - 1] is None:
                continue
            values[col[row_num1 - 1]] = col[row_num2 - 1]
        return values

    # 获取指定行，组成列表
    def rows_values_to_list(self, sheet_name, row_num):
        sheet = self.workbook[sheet_name]
        values = []
        for col in sheet.iter_cols(min_row=1, values_only=True):
            # 如果为None，跳过此次
            if col[row_num - 1] is None:
                continue
            values.append(col[row_num - 1])
        return values

    # 读取excel所有数据
    def read_excel_all(self, sheet_name):
        # 打开一个现有的Excel文件
        sheet = self.workbook[sheet_name]
        # 获取活动的工作表
        # 存储所有行数据的列表
        data_list = []
        # 迭代所有行和列，将每行数据添加到列表中
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index == 0:
                continue  # 跳过第一行
            data_list.append(list(row))
        return data_list

    # 将列表list写入指定列
    def write_list_to_column(self, sheet_name, col_num, data_list):
        sheet = self.workbook[sheet_name]
        if not isinstance(data_list, list):
            raise TypeError(f"'{data_list}' must be a list")
        for i in range(len(data_list)):
            # 在写入数据时，从第二行开始写入，因为第一行通常是标题行。
            sheet.cell(row=i + 2, column=col_num, value=data_list[i])
        self.save()

    # 将列表list写入指定行
    def write_list_to_row(self, sheet_name, row_num, data_list):
        if not isinstance(data_list, list):
            raise TypeError(f"'{data_list}' must be a list")
        sheet = self.workbook[sheet_name]

        col_num = 1
        for item in data_list:
            sheet.cell(row=row_num, column=col_num, value=item)
            col_num += 1
        self.save()
    # 将字典的键和值，写入指定列中
    def write_dict_to_col(self, sheet_name, dict_data, col1, col2):
        if not isinstance(dict_data, dict):
            raise TypeError(f"'{dict_data}' must be a dict")
        sheet = self.workbook[sheet_name]
        # 在写入数据时，从第二行开始写入，因为第一行通常是标题行。
        row_num = 2
        for key, value in dict_data.items():
            sheet.cell(row=row_num, column=col1, value=key)
            sheet.cell(row=row_num, column=col2, value=value)
            row_num += 1
        self.save()

    # 将字典的键和值，写入指定行中
    def write_dict_to_row(self, sheet_name, dict_data, row1, row2):
        if not isinstance(dict_data, dict):
            raise TypeError(f"'{dict_data}' must be a dict")
        sheet = self.workbook[sheet_name]
        col_num = 1
        for key, value in dict_data.items():
            sheet.cell(row=row1, column=col_num, value=key)
            sheet.cell(row=row2, column=col_num, value=value)
            col_num += 1
        self.save()



if __name__ == '__main__':
    filepath = os.path.join(config.DATA_DIR, 'FT Network List - Updated on 20240617.xlsx')
    excel_obj = ExcelFile(filepath)
    data=excel_obj.column_values_to_set("工作表1",2)
    print(list(data))